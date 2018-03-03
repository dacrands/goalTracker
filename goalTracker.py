#! python3.
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 15 16:19:20 2017

@author: dacrands
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date

#TODO prompt user to create goals
try:
    wb = load_workbook(filename = "goal-tracker.xlsx")
except IOError:
    print("Creating workbook")
    wb = Workbook()
    
    # Create sheet names (meditate, walk, read)
    walkSheet = wb.active
    walkSheet.title = "walk"
    walkSheet.cell(row=1, column=1).value  = "DATE"
    walkSheet.cell(row=1, column=2).value  = "MINUTES"
                  
    medSheet = wb.create_sheet(title="meditate")
    medSheet.cell(row=1, column=1).value  = "DATE"
    medSheet.cell(row=1, column=2).value  = "MINUTES"
                 
    readSheet = wb.create_sheet(title="read")
    readSheet.cell(row=1, column=1).value = "DATE"
    readSheet.cell(row=1, column=2).value = "MINUTES"
                  
    wb.save("goal-tracker.xlsx")
    
def newSheet():
    leaveNew = False
    print("What activity would you like to add:")
    name = input(">>> ")
    while not leaveNew:
        try:
            newSheet = wb.create_sheet(name)
            newSheet.cell(row=1, column=1).value  = "DATE"
            newSheet.cell(row=1, column=2).value  = "MINUTES"
            leaveNew = True
        except ValueError:
            print("invalid");

    

def chooseSheet():
    leaveSheet = False
    while not leaveSheet:
        print("Choose an activity: ")
        for i in range(len(wb.sheetnames)):
            print("{0}: {1}\r".format(i, wb.sheetnames[i]))
            
        sheetChoice = input(">>>")        
        try:
            sheetChoice = int(sheetChoice)
            currSheet = wb.sheetnames[sheetChoice]
            print(currSheet)
            leaveSheet = True
            return currSheet 
           
        except:
            print("That's not a valid option")


leave = False
while not leave:
    print("Press 'n' to update goal.\nPress 'a' to add a goal.\nPress 'q' to quit")
    choices = ['q', 'n', 'a']
    choice = input(">>>")
        
    try:
        if choice not in choices:
            raise ValueError()
            
        elif choice == 'q':
            leave = True
            
        elif choice == 'a':
            newSheet()
            
        elif choice == 'n':
            sheetLeave = False
            while not sheetLeave:
                currSheet = chooseSheet()
                sheet = wb[currSheet]   
                print(sheet.max_row)
                time = input("How long were you doing said task? >>")
                
                try: 
                    int(time)
                    for rowNum in range(1, sheet.max_row+2):
                        if sheet.cell(row=rowNum, column=1).value is None:
                            sheet.cell(row=rowNum, column=1).value = str(date.today())
                            sheet.cell(row=rowNum, column=2).value = time
                            print("You successfully updated {0}".format(currSheet))
                            sheetLeave = True
                            
                except ValueError:
                    print("Please enter an integer, representing minutes...REPRESENT, B!")
                
    except ValueError:
        print("That's not a valid option")


wb.save("goal-tracker.xlsx")


